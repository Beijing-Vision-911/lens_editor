import sys
from functools import partial
from itertools import chain
from pathlib import Path

from openpyxl import Workbook
from PySide6.QtCore import QMutex, QThreadPool,Qt
from PySide6.QtGui import QKeySequence, QPixmapCache, QShortcut,QBrush, QColor
from PySide6.QtWidgets import (QApplication, QCompleter, QFileDialog,
                               QGraphicsGridLayout, QGraphicsScene,
                               QGraphicsWidget, QHBoxLayout,
                               QInputDialog, QLineEdit, QMainWindow,
                               QMessageBox, QPushButton, QStatusBar,
                               QVBoxLayout, QWidget)

from .config import root_config
from .defect import DefectItem, Lens,DefectEdit
from .rule import linemapping, xymapping
from .rule_edit import RuleEditWindow
from .search import FilterParser, QuickSearchSlot
from .thread import Worker
from .view import View
import logging


level = logging.ERROR if sys.argv[-1] != "-d" else logging.INFO
logging.basicConfig(level=level, format="%(levelname)s:%(message)s")
logger = logging.getLogger(__name__)


class MainWindow(QMainWindow):
    def __init__(self, initial_path=""):
        super().__init__()
        widget = QWidget()
        main_layout = QVBoxLayout()
        self.scene = QGraphicsScene()
        self.main_view = View(self.scene)
        QPixmapCache.setCacheLimit(1024 * 1024 * 10)
        self.num = 0
        self.scene.setItemIndexMethod(QGraphicsScene.NoIndex)
        main_layout.addWidget(self.main_view)

        bottom_layout = QHBoxLayout()
        main_layout.addLayout(bottom_layout)
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        self.search_bar = QLineEdit()
        self.search_bar.returnPressed.connect(
            lambda: self.filter_apply(self.search_bar.text())
        )
        bottom_layout.addWidget(self.search_bar)

        self.rule_edit_btn = QPushButton("Rule")
        self.rule_edit_btn.clicked.connect(self.rule_edit_btn_clicked)
        bottom_layout.addWidget(self.rule_edit_btn)

        self.mark_btn = QPushButton("Mark(a)")
        self.mark_btn.clicked.connect(self.mark_btn_clicked)
        bottom_layout.addWidget(self.mark_btn)

        self.save_btn = QPushButton("Save(s)")
        self.save_btn.clicked.connect(self.save_btn_clicked)
        bottom_layout.addWidget(self.save_btn)

        self.rename_btn = QPushButton("Rename(r)")
        self.rename_btn.clicked.connect(self.rename_btn_clicked)
        bottom_layout.addWidget(self.rename_btn)

        self.open_file = QPushButton("Open(o)")
        self.open_file.clicked.connect(self.btn_openfile)
        bottom_layout.addWidget(self.open_file)

        self.count_bar = QLineEdit()
        bottom_layout.addWidget(self.count_bar)

        self.count_btn = QPushButton("Count(c)")
        self.count_btn.clicked.connect(self.count_btn_clicked)
        bottom_layout.addWidget(self.count_btn)

        self.export_btn = QPushButton("Export")
        bottom_layout.addWidget(self.export_btn)

        self.delete_btn = QPushButton("Delete")
        self.delete_btn.clicked.connect(self.delete)
        bottom_layout.addWidget(self.delete_btn)

        widget.setLayout(main_layout)
        self.setWindowTitle("Lens Editor")
        self.setGeometry(0, 0, 1000, 800)
        self.setCentralWidget(widget)

        self.thread_pool = QThreadPool()
        self.mutex = QMutex()
        self.filter_parser = FilterParser()

        self.shortcuts()

        if initial_path:
            self._load_files(initial_path)

    def shortcuts(self):
        QShortcut(QKeySequence("o"), self, self.btn_openfile)
        QShortcut(QKeySequence("s"), self, self.save_btn_clicked)
        QShortcut(QKeySequence("a"), self, self.mark_btn_clicked)
        QShortcut(QKeySequence("r"), self, self.rename_btn_clicked)
        QShortcut(QKeySequence("c"), self, self.count_btn_clicked)

        self.search_slot = QuickSearchSlot()

        slot_apply = lambda i: self.filter_apply(self.search_slot.get_slot(i), True)
        slot_set = lambda i: self.search_slot.set_slot(i, self.search_bar.text())
        for i in map(str, range(1, 9)):
            QShortcut(i, self, partial(slot_apply, i))
            QShortcut(QKeySequence(f"Ctrl+{i}"), self, partial(slot_set, i))



    def delete(self):
        if not hasattr(self, "scene"):
            return
        items = self.scene.selectedItems()
        if len(items) == 0:
            self.status_bar.showMessage("No item selected")
            return
        else:
            for i in items:
                i.delete_xml()
            self.status_bar.showMessage(" Delete succeeded  ")
            



    def rule_edit_btn_clicked(self):
        self.rule_window = RuleEditWindow(main_window=self)
        self.export_btn.clicked.connect(self.rule_window.export_btn_clicked)
        self.rule_window.show()

    def rename_btn_clicked(self):
        if not hasattr(self, "scene"):
            return
        items = self.scene.selectedItems()
        if len(items) == 0:
            self.status_bar.showMessage("No item selected")
            return
        new_label, ok = QInputDialog.getText(self, "Rename", "New Label:")
        if not ok:
            return

        for i in items:
            i.rename(new_label)

    def save_btn_clicked(self):
        mod_files_num = len([i for i in self.lens if i.modified])
        for i in self.lens:
            i.save()
        self.status_bar.showMessage(f"Saved {mod_files_num} changes")

    def mark_btn_clicked(self):
        if not hasattr(self, "scene"):
            return
        items = self.scene.selectedItems()
        mark_state = [i.mark_toggle() for i in items]
        marked = len([x for x in mark_state if x])
        unmarked = len([x for x in mark_state if not x])
        message = ""
        if marked > 0:
            message += f"Marked {marked} items. "
        if unmarked > 0:
            message += f"Unmarked {unmarked} items."

        self.status_bar.showMessage(message)

    def filter_apply(self, query, search_bar_update=False):

        if query =="A":
            print("1")
        # elif query[-1] == "A":
        #     pass
        d_list = self.filter_parser.parse(query, self.defects)
        self.view_update(d_list)
        self.status_bar.showMessage(
            f"Filter: {self.search_bar.text()}, Total: {len(d_list)}"
        )
        if search_bar_update:
            self.search_bar.setText(query)

    def _load_files(self, path: str) -> None:
        xml_files = [x for x in Path(path).glob("**/*.xml") if x.is_file()]

        def find_jpeg(xml_file):
            if (jpeg_file := xml_file.with_suffix(".jpeg")).is_file():
                return jpeg_file
            if (
                jpeg_file := xml_file.parents[1] / "img" / f"{xml_file.stem}.jpeg"
            ).is_file():
                return jpeg_file
            if (jpeg_file := xml_file.with_suffix(".jpg")).is_file():
                return jpeg_file
            logger.error(f"Cannot find jpeg for {xml_file}")
            return None

        parms = [(x, find_jpeg(x)) for x in xml_files if find_jpeg(x)]

        self.lens = []
        self.total_file = len(parms)
        self.processed_file = 0
        for f, j in parms:
            w = Worker(Lens, f, j)
            w.signals.result.connect(self.worker_done)
            self.thread_pool.start(w)

    def btn_openfile(self):
        file_path = QFileDialog.getExistingDirectory()
        self._load_files(file_path)


    def worker_done(self, lz):
        self.mutex.lock()
        self.lens.append(lz)
        self.processed_file += 1
        self.status_bar.showMessage(
            f"Loading files: {self.total_file}/{self.processed_file}"
        )
        self.mutex.unlock()

        if self.processed_file == self.total_file:
            self.defects = sorted(
                chain(*[l.defects for l in self.lens]),
                key=lambda x: (x.name, x.width, x.height),
            )
            self.view_update(self.defects)
            complete_candidates = list(set([d.name for d in self.defects]))
            completer = QCompleter(complete_candidates)
            self.search_bar.setCompleter(completer)
            self.status_bar.showMessage(
                f"No Filter, Category: {len(complete_candidates)},Total: {len(self.defects)}"
            )

    def view_update(self, d_list):
        g_layout = QGraphicsGridLayout()
        g_layout.setContentsMargins(10, 10, 10, 10)
        g_layout.setSpacing(25)
        g_widget = QGraphicsWidget()
        self.scene.clear()
        # dynamic column size, dependents on window width
        col_size = int(self.frameGeometry().width() / 80)
        for i, di in enumerate([DefectItem(d).get_layout_item() for d in d_list]):
            r = int(i / col_size)
            c = i % col_size
            g_layout.addItem(di, r, c)
        g_widget.setLayout(g_layout)
        self.scene.addItem(g_widget)
        self.main_view.centerOn(self.scene.itemsBoundingRect().center())


    def keyPressEvent(self, event):
        self.num+=1
        if event.key() == Qt.Key_Space:
            if not hasattr(self, "scene"):
                return
            items = self.scene.selectedItems()
            if len(items) == 0:
                QMessageBox.information(self,"丢失","目标丢失,尝试按“Ctrl+z”以重新定位缺陷")
                self.status_bar.showMessage("No item selected")
                return
            elif len(items) == 1:
                if self.num == 1:
                    QMessageBox.information(self,"连续模式","已经进入连续模式,按下”space“以切换下一张缺陷")
                # QMessageBox.about(self,)

                index = self.defects.index(items[0].defect)    #defect 的索引  0......+1
                self.index1 = self.scene.items().index(items[0])    # items的索引   1.....+3
                self.scene.clearSelection()
                item = self.scene.items()[self.index1+3]      #选中下一个item
                item.setSelected(True)
                item.change_color()
                
                try:
                    defect_edit = DefectEdit(self.defects[index+1])
                    defect_edit.edit()
                except IndexError:
                    self.status_bar.showMessage("当前是最后一张")
            else:
                self.status_bar.showMessage("Too Many Items")
                return
        elif event.key() == Qt.Key_Z and event.modifiers() == Qt.ControlModifier:
            item = self.scene.items()[self.index1]
            item.setSelected(True)
            item.change_color()
            self.status_bar.showMessage("已选上一个缺陷，按下”space“以打开缺陷")


    def count_btn_clicked(self):
        text = self.count_bar.text()
        li = ["0", "1", "2", "3", "4"]
        nums = []
        names = []
        w = []
        h = []
        regions = []
        recs = []
        lnames = []
        wb = Workbook()
        ws = wb.active
        ws.title = "sheet"
        ws.append(["编号", "类名", "宽度", "高度", "区域", "是否可收", "左通道对应"])
        for d in self.defects:
            if d.name == text:
                receive = str(d.lens.xml_path)
                r = receive.split("/")
                if "unq" in r or "UNQ" in r:
                    rec = "UNQ"
                elif "h1" in r or "H1" in r:
                    rec = "H1"
                elif "h2" in r or "H2" in r:
                    rec = "H2"
                else:
                    rec = ""
                recs.append(rec)
                img_number = d.lens.xml_path.stem
                nums.append(img_number)
                names.append(d.name)
                w.append(d.width)
                h.append(d.height)
                x = d.xmin
                if x > 1329 and x < 1711:
                    region = "A"
                elif x > 1710 and x < 1961:
                    region = "B"
                elif x > 1960 and x < 2210:
                    region = "C"
                else:
                    region = "D"
                regions.append(region)
                ln = []
                for i in li:
                    if d.name[0] == "0" or d.name[0] == "1":
                        a = self.left_check(d, i)
                    else:
                        a = self.line_check(d, i)
                        if a:
                            ln.append(a)
                lnames.append(str(ln))
        for i in range(len(nums)):
            ws.cell(i + 2, 1, nums[i])
            ws.cell(i + 2, 2, names[i])
            ws.cell(i + 2, 3, w[i])
            ws.cell(i + 2, 4, h[i])
            ws.cell(i + 2, 5, regions[i])
            ws.cell(i + 2, 6, recs[i])
            ws.cell(i + 2, 7, lnames[i])
        wb.save(f"{text}.xlsx")

    def left_check(self, d, sexp):
        fn = xymapping(d.x, d.y)
        mappings = [d_.name for d_ in d.lens.left if fn(d_.x, d_.y)]
        for name in mappings:
            if name.endswith(sexp):
                return name
        return False

    def line_check(self, d, sexp):
        fn = linemapping(d.x, d.y, d.x_, d.y_)
        mappings = [d_.name for d_ in d.lens.left if fn(d_.x, d_.y, d_.x_, d_.y_)]
        for name in mappings:
            if name.endswith(sexp):
                return name
        return False

    def closeEvent(self, event) -> None:
        reply = QMessageBox.question(
            self, "提示", "是否关闭所有窗口", QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            event.accept()
            sys.exit(0)
        else:
            event.ignore()


def main():
    root_config("~/.lens_editor")
    app = QApplication(sys.argv)
    initial_path = sys.argv[1] if len(sys.argv) > 1 else ""
    window = MainWindow(initial_path)
    window.show()
    app.exec()
